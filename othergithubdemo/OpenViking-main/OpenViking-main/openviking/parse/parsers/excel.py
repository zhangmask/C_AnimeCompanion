# Copyright (c) 2026 Beijing Volcano Engine Technology Co., Ltd.
# SPDX-License-Identifier: AGPL-3.0
"""
Excel (.xlsx/.xls/.xlsm) parser for OpenViking.

Converts Excel spreadsheets to Markdown then parses using MarkdownParser.
Inspired by microsoft/markitdown approach.
"""

import asyncio
from pathlib import Path
from typing import List, Optional, Union

from openviking.parse.base import ParseResult
from openviking.parse.parsers.base_parser import BaseParser
from openviking_cli.utils.config.parser_config import ParserConfig
from openviking_cli.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelParser(BaseParser):
    """
    Excel spreadsheet parser for OpenViking.

    Supports: .xlsx, .xls, .xlsm

    Converts Excel spreadsheets to Markdown using openpyxl,
    then delegates to MarkdownParser for tree structure creation.
    """

    def __init__(self, config: Optional[ParserConfig] = None, max_rows_per_sheet: int = 1000):
        """
        Initialize Excel parser.

        Args:
            config: Parser configuration
            max_rows_per_sheet: Maximum rows to process per sheet (0 = unlimited)
        """
        from openviking.parse.parsers.markdown import MarkdownParser

        self._md_parser = MarkdownParser(config=config)
        self.config = config or ParserConfig()
        self.max_rows_per_sheet = max_rows_per_sheet

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls", ".xlsm"]

    async def parse(self, source: Union[str, Path], instruction: str = "", **kwargs) -> ParseResult:
        """Parse Excel spreadsheet from file path."""
        path = Path(source)

        if path.exists():
            # Use xlrd for legacy .xls, openpyxl for .xlsx/.xlsm
            if path.suffix.lower() == ".xls":
                markdown_content = await asyncio.to_thread(self._convert_xls_to_markdown, path)
            else:
                import openpyxl

                markdown_content = await asyncio.to_thread(
                    self._convert_to_markdown, path, openpyxl
                )
            result = await self._md_parser.parse_content(
                markdown_content, source_path=str(path), instruction=instruction, **kwargs
            )
        else:
            result = await self._md_parser.parse_content(
                str(source), instruction=instruction, **kwargs
            )
        result.source_format = path.suffix.lstrip(".") if path.exists() else "xlsx"
        result.parser_name = "ExcelParser"
        return result

    async def parse_content(
        self, content: str, source_path: Optional[str] = None, instruction: str = "", **kwargs
    ) -> ParseResult:
        """Parse content - delegates to MarkdownParser."""
        result = await self._md_parser.parse_content(content, source_path, **kwargs)
        result.source_format = "xlsx"
        result.parser_name = "ExcelParser"
        return result

    def _convert_xls_to_markdown(self, path: Path) -> str:
        """Convert legacy .xls spreadsheet to Markdown using xlrd."""
        import xlrd

        # formatting_info=True enables xlrd to detect date cells via XL_CELL_DATE
        # instead of reporting them as XL_CELL_NUMBER with raw float serials
        wb = xlrd.open_workbook(str(path), formatting_info=True, on_demand=True)
        try:
            return self._build_xls_markdown(wb, path, xlrd)
        finally:
            wb.release_resources()

    def _build_xls_markdown(self, wb, path: Path, xlrd) -> str:
        """Build markdown from xlrd workbook."""
        markdown_parts = []
        markdown_parts.append(f"# {path.stem}")
        markdown_parts.append(f"**Sheets:** {wb.nsheets}")

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            parts = [f"## Sheet: {sheet.name}"]

            if sheet.nrows == 0 or sheet.ncols == 0:
                parts.append("*Empty sheet*")
                markdown_parts.append("\n\n".join(parts))
                continue

            parts.append(f"**Dimensions:** {sheet.nrows} rows × {sheet.ncols} columns")

            rows_to_process = sheet.nrows
            if self.max_rows_per_sheet > 0:
                rows_to_process = min(sheet.nrows, self.max_rows_per_sheet)

            rows = []
            for row_idx in range(rows_to_process):
                row_data = []
                for col_idx in range(sheet.ncols):
                    row_data.append(self._format_xls_cell(sheet.cell(row_idx, col_idx), wb, xlrd))
                rows.append(row_data)

            if rows:
                from openviking.parse.base import format_table_to_markdown

                parts.append(format_table_to_markdown(rows, has_header=True))

            if self.max_rows_per_sheet > 0 and sheet.nrows > self.max_rows_per_sheet:
                parts.append(
                    f"\n*... {sheet.nrows - self.max_rows_per_sheet} more rows truncated ...*"
                )

            markdown_parts.append("\n\n".join(parts))

        return "\n\n".join(markdown_parts)

    @staticmethod
    def _format_xls_cell(cell, wb, xlrd) -> str:
        """Format a single xlrd cell value with proper type handling."""
        if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                # Include time component if non-zero
                if dt[3] or dt[4] or dt[5]:
                    return (
                        f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}"
                    )
                return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}"
            except Exception:
                return str(cell.value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell.value else "FALSE"
        if cell.ctype == xlrd.XL_CELL_ERROR:
            # xlrd error code map
            error_map = {
                0x00: "#NULL!",
                0x07: "#DIV/0!",
                0x0F: "#VALUE!",
                0x17: "#REF!",
                0x1D: "#NAME?",
                0x24: "#NUM!",
                0x2A: "#N/A",
            }
            return error_map.get(cell.value, f"#ERR({cell.value})")
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            # Display integers without trailing .0
            if cell.value == int(cell.value):
                return str(int(cell.value))
            return str(cell.value)
        # XL_CELL_TEXT or fallback
        return str(cell.value) if cell.value is not None else ""

    def _convert_to_markdown(self, path: Path, openpyxl) -> str:
        """Convert Excel spreadsheet to Markdown string."""
        wb = openpyxl.load_workbook(path, data_only=True)

        markdown_parts = []
        markdown_parts.append(f"# {path.stem}")
        markdown_parts.append(f"**Sheets:** {len(wb.sheetnames)}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_content = self._convert_sheet(sheet, sheet_name)
            markdown_parts.append(sheet_content)

        return "\n\n".join(markdown_parts)

    def _convert_sheet(self, sheet, sheet_name: str) -> str:
        """Convert a single sheet to markdown."""
        parts = []
        parts.append(f"## Sheet: {sheet_name}")

        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            parts.append("*Empty sheet*")
            return "\n\n".join(parts)

        parts.append(f"**Dimensions:** {max_row} rows × {max_col} columns")

        rows_to_process = max_row
        if self.max_rows_per_sheet > 0:
            rows_to_process = min(max_row, self.max_rows_per_sheet)

        rows = []
        for _row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=rows_to_process, values_only=True), 1
        ):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell))
            rows.append(row_data)

        if rows:
            from openviking.parse.base import format_table_to_markdown

            table_md = format_table_to_markdown(rows, has_header=True)
            parts.append(table_md)

        if self.max_rows_per_sheet > 0 and max_row > self.max_rows_per_sheet:
            parts.append(f"\n*... {max_row - self.max_rows_per_sheet} more rows truncated ...*")

        return "\n\n".join(parts)
